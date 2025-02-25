import pandas as pd
import openpyxl
import os
import numpy as np
import re
import json
import traceback
import math


def transform_excel_file(input_file="input.xlsx", output_file="output.xlsx", exception_file="exceptions.json"):
    """
    1번 Excel 파일을 2번 파일과 같은 형식으로 변환하는 함수
    예외 상품은 장수와 무관하게 기본 상품명으로 비교하여 처리
    """
    try:
        print(f"파일 로딩 중: {input_file}")
        
        # 먼저 pandas로 데이터 로드
        df = pd.read_excel(input_file, sheet_name=0, dtype=str)
        
        # 예외 목록 로드
        exception_products = []
        sheet_limits = {}  # 상품별 장수 한계 저장
        
        if os.path.exists(exception_file):
            try:
                with open(exception_file, 'r', encoding='utf-8') as f:
                    exception_data = json.load(f)
                    if "exception_products" in exception_data:
                        exception_products = exception_data["exception_products"]
                        
                        # 장수 한계 추출 및 저장
                        for product in exception_products:
                            try:
                                match = re.search(r'(.+?)\s+(\d+)장$', product)
                                if match:
                                    base_name = match.group(1).strip()
                                    sheet_limit = int(match.group(2))
                                    sheet_limits[base_name] = sheet_limit
                                    print(f"장수 한계 설정: {base_name} - {sheet_limit}장")
                            except Exception as e:
                                print(f"장수 한계 추출 오류: {product}, {str(e)}")
                        
                print(f"예외 상품 목록 로드 완료: {len(exception_products)}개 항목")
                print(f"예외 상품: {exception_products}")
            except Exception as e:
                print(f"예외 목록 로드 중 오류 발생: {str(e)}")
                print("예외 없이 계속 진행합니다.")
        else:
            print(f"예외 파일 '{exception_file}'이 존재하지 않습니다. 예외 없이 계속 진행합니다.")
        
        required_columns = ['주문번호', '상태', '상품명-옵션명', '관리용상품명', '수량', 
                           '받는분', '받는분 연락처', '배송지 우편번호', '도로명 주소', '배송메시지']
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            print(f"오류: 필요한 열이 없습니다: {', '.join(missing_columns)}")
            return False
        
        # NaN 값 처리
        df = df.replace({np.nan: ''})
        
        # 1. 데이터 전처리 - 필요한 열만 선택
        required_data = df[[
            '주문번호', '상태', '상품명-옵션명', '관리용상품명', '수량', 
            '받는분', '받는분 연락처', '배송지 우편번호', '도로명 주소', '배송메시지'
        ]].copy()
        
        # 빈 행 제거
        required_data = required_data[required_data['주문번호'] != '']
        
        # 2. 주문 데이터를 리스트로 변환 (행별 처리용)
        rows = []
        for _, row in required_data.iterrows():
            try:
                # 필요한 데이터 추출
                customer_info = {
                    'name': row['받는분'],
                    'phone': row['받는분 연락처'],
                    'postal_code': row['배송지 우편번호'],
                    'address': row['도로명 주소'],
                    'message': row['배송메시지']
                }
                
                product_name = str(row['관리용상품명']).strip()
                quantity_str = str(row['수량']).strip()
                quantity = int(quantity_str) if quantity_str and quantity_str.isdigit() else 1
                
                # 장수 패턴 추출
                base_product_name = product_name
                sheet_count = 0
                
                try:
                    match = re.search(r'(.+?)\s+(\d+)장$', product_name)
                    if match:
                        base_product_name = match.group(1).strip()
                        sheet_count = int(match.group(2))
                except Exception as e:
                    print(f"장수 패턴 추출 중 오류: {product_name}, {str(e)}")
                
                # 예외 상품인지 확인 - 정확히 일치하는 경우
                is_exception = product_name in exception_products
                
                rows.append({
                    'customer_info': customer_info,
                    'product': product_name,
                    'base_product': base_product_name,
                    'sheet_count': sheet_count,
                    'quantity': quantity,
                    'is_exception': is_exception,
                    'order_num': row['주문번호']  # 디버깅용으로 주문번호 추가
                })
            except Exception as e:
                print(f"행 처리 중 오류: {str(e)}")
                continue
        
        # 3. 행별 순차 비교 및 병합
        merged_rows = []
        skip_indices = set()  # 이미 병합된 행 인덱스
        
        for i in range(len(rows)):
            if i in skip_indices:
                continue  # 이미 병합된 행은 건너뛰기
            
            current_row = rows[i]
            
            # 예외 상품인 경우 - 개별 처리
            if current_row['is_exception']:
                print(f"예외 상품 처리 중: {current_row['product']}, 수량: {current_row['quantity']}")
                
                # 예외 상품은 그대로 주문 수량만큼 행 추가
                for _ in range(current_row['quantity']):
                    exception_row = [
                        current_row['customer_info']['name'],
                        current_row['customer_info']['phone'],
                        current_row['customer_info']['postal_code'],
                        current_row['customer_info']['address'],
                        current_row['customer_info']['message'],
                        current_row['product'],  # 원본 상품명 그대로 사용
                        "1"  # 수량은 항상 1
                    ]
                    merged_rows.append(exception_row)
            else:
                try:
                    # 일반 상품은 순차 비교하여 병합
                    customer_key = "_".join([
                        current_row['customer_info']['name'],
                        current_row['customer_info']['phone'],
                        current_row['customer_info']['postal_code'],
                        current_row['customer_info']['address'],
                        current_row['customer_info']['message']
                    ])
                    
                    # 현재 행의 병합 대상 상품들
                    merged_products = []  # 상품명 
                    merged_quantities = []  # 수량
                    merged_base_products = []  # 기본 상품명
                    merged_sheet_counts = []  # 장수 정보
                    
                    # 현재 상품 추가
                    merged_products.append(current_row['product'])
                    merged_quantities.append(current_row['quantity'])
                    merged_base_products.append(current_row['base_product'])
                    merged_sheet_counts.append(current_row['sheet_count'])
                    
                    # 다음 행부터 순차적으로 비교
                    for j in range(i + 1, len(rows)):
                        if j in skip_indices:
                            continue  # 이미 병합된 행은 건너뛰기
                        
                        next_row = rows[j]
                        
                        # 예외 상품인 경우 병합하지 않고 건너뜀
                        if next_row['is_exception']:
                            continue
                        
                        next_customer_key = "_".join([
                            next_row['customer_info']['name'],
                            next_row['customer_info']['phone'],
                            next_row['customer_info']['postal_code'],
                            next_row['customer_info']['address'],
                            next_row['customer_info']['message']
                        ])
                        
                        # 고객 정보가 같으면 병합 가능성 검토
                        if customer_key == next_customer_key:
                            try:
                                # 동일한 상품인지 확인
                                same_product_found = False
                                should_skip_merge = False  # 장수 제한으로 병합을 건너뛸지 여부
                                
                                for k in range(len(merged_products)):
                                    product_str = merged_products[k]
                                    base_product = merged_base_products[k]
                                    
                                    # 상품명이 정확히 일치하는 경우 수량만 더함
                                    if product_str == next_row['product']:
                                        # 장수 제한 확인
                                        if base_product in sheet_limits:
                                            limit = sheet_limits[base_product]
                                            current_sheets = merged_sheet_counts[k] * merged_quantities[k]
                                            next_sheets = next_row['sheet_count'] * next_row['quantity']
                                            total_sheets = current_sheets + next_sheets
                                            
                                            if total_sheets > limit:
                                                print(f"장수 제한 초과: {base_product}의 장수가 {total_sheets}장으로 {limit}장을 초과")
                                                should_skip_merge = True
                                                break
                                        
                                        if not should_skip_merge:
                                            merged_quantities[k] += next_row['quantity']
                                            same_product_found = True
                                        break
                                    
                                    # 기본 상품명이 같은 경우도 병합
                                    elif base_product == next_row['base_product']:
                                        # 장수 제한 확인
                                        if base_product in sheet_limits:
                                            limit = sheet_limits[base_product]
                                            # 현재 장수 계산
                                            current_sheets = 0
                                            if merged_sheet_counts[k] > 0:
                                                current_sheets = merged_sheet_counts[k] * merged_quantities[k]
                                            else:
                                                current_sheets = merged_quantities[k]
                                            
                                            # 다음 장수 계산
                                            next_sheets = 0
                                            if next_row['sheet_count'] > 0:
                                                next_sheets = next_row['sheet_count'] * next_row['quantity']
                                            else:
                                                next_sheets = next_row['quantity']
                                            
                                            total_sheets = current_sheets + next_sheets
                                            
                                            if total_sheets > limit:
                                                print(f"장수 제한 초과: {base_product}의 장수가 {total_sheets}장으로 {limit}장을 초과")
                                                should_skip_merge = True
                                                break
                                        
                                        if should_skip_merge:
                                            break
                                            
                                        # 장수 형식 처리
                                        existing_match = re.search(r'(.+?)\s+(\d+)장$', product_str)
                                        new_match = re.search(r'(.+?)\s+(\d+)장$', next_row['product'])
                                        
                                        if existing_match and new_match:
                                            # 둘 다 장수가 있는 경우
                                            existing_sheets = int(existing_match.group(2))
                                            new_sheets = int(new_match.group(2))
                                            
                                            # 장수와 수량 고려하여 합산
                                            total_sheets = (existing_sheets * merged_quantities[k]) + (new_sheets * next_row['quantity'])
                                            merged_products[k] = f"{base_product} {total_sheets}장"
                                            merged_quantities[k] = 1  # 이미 장수에 수량 반영됨
                                            merged_sheet_counts[k] = total_sheets
                                            same_product_found = True
                                            break
                                        elif existing_match:
                                            # 기존 상품에만 장수가 있는 경우
                                            existing_sheets = int(existing_match.group(2))
                                            total_quantity = merged_quantities[k] + next_row['quantity']
                                            total_sheets = existing_sheets * total_quantity
                                            merged_products[k] = f"{base_product} {total_sheets}장"
                                            merged_quantities[k] = 1  # 이미 장수에 수량 반영됨
                                            merged_sheet_counts[k] = total_sheets
                                            same_product_found = True
                                            break
                                        elif new_match:
                                            # 새 상품에만 장수가 있는 경우
                                            new_sheets = int(new_match.group(2))
                                            total_quantity = merged_quantities[k] + next_row['quantity']
                                            total_sheets = new_sheets * total_quantity
                                            merged_products[k] = f"{base_product} {total_sheets}장"
                                            merged_quantities[k] = 1  # 이미 장수에 수량 반영됨
                                            merged_sheet_counts[k] = total_sheets
                                            same_product_found = True
                                            break
                                        else:
                                            # 둘 다 장수가 없는 경우는 수량만 합산
                                            merged_quantities[k] += next_row['quantity']
                                            same_product_found = True
                                            break
                                
                                # 장수 제한으로 건너뛰어야 하는 경우
                                if should_skip_merge:
                                    continue
                                
                                # 같은 상품을 찾지 못했으면 새로 추가
                                if not same_product_found:
                                    merged_products.append(next_row['product'])
                                    merged_quantities.append(next_row['quantity'])
                                    merged_base_products.append(next_row['base_product'])
                                    merged_sheet_counts.append(next_row['sheet_count'])
                                
                                skip_indices.add(j)  # 사용된 행 표시
                            except Exception as e:
                                print(f"상품 병합 중 오류: {str(e)}")
                                traceback.print_exc()
                        else:
                            # 고객 정보가 다르면 더이상 비교하지 않고 다음 행으로 넘어감
                            break
                    
                    # 상품 정보와 수량을 결합하여 최종 상품 리스트 생성
                    final_products = []
                    for prod, qty, sheet_count in zip(merged_products, merged_quantities, merged_sheet_counts):
                        try:
                            if sheet_count > 0 and qty > 1:
                                # 장수 형식이면 장수와 수량을 곱하여 총 장수 계산
                                match = re.search(r'(.+?)\s+(\d+)장$', prod)
                                if match:
                                    base_name = match.group(1).strip()
                                    total_sheets = sheet_count * qty
                                    final_products.append(f"{base_name} {total_sheets}장")
                                else:
                                    final_products.append(f"{prod} {qty}장")
                            elif qty > 1:
                                # 장수 형식이 아니면 수량을 붙여서 표시
                                final_products.append(f"{prod} {qty}장")
                            else:
                                # 수량이 1이면 그대로 사용
                                final_products.append(prod)
                        except Exception as e:
                            print(f"상품 형식 처리 중 오류: {prod}, {str(e)}")
                            final_products.append(prod)  # 오류 시 원본 사용
                    
                    # 일반 상품 병합하여 한 행으로 추가
                    products_str = " ,".join(final_products)
                    merged_row = [
                        current_row['customer_info']['name'],
                        current_row['customer_info']['phone'],
                        current_row['customer_info']['postal_code'],
                        current_row['customer_info']['address'],
                        current_row['customer_info']['message'],
                        products_str,
                        "1"  # 최종 수량은 항상 1
                    ]
                    merged_rows.append(merged_row)
                except Exception as e:
                    print(f"행 병합 중 오류: {str(e)}")
                    traceback.print_exc()
                    # 오류 발생 시 원본 행을 그대로 추가
                    merged_row = [
                        current_row['customer_info']['name'],
                        current_row['customer_info']['phone'],
                        current_row['customer_info']['postal_code'],
                        current_row['customer_info']['address'],
                        current_row['customer_info']['message'],
                        current_row['product'],
                        str(current_row['quantity'])
                    ]
                    merged_rows.append(merged_row)
        
        # 데이터가 없는 경우 처리
        if not merged_rows:
            print("변환할 데이터가 없습니다.")
            return False
        
        # 새 워크북 생성
        new_workbook = openpyxl.Workbook()
        new_sheet = new_workbook.active
        new_sheet.title = "주문관리목록"  # 첫 번째 시트 이름 설정
        
        # 변환된 데이터 입력 - 모든 셀을 문자열로 저장
        for row_idx, row_data in enumerate(merged_rows, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                cell = new_sheet.cell(row=row_idx, column=col_idx)
                cell.value = cell_value
                
                # 전화번호(B열)는 문자열 서식 적용
                if col_idx == 2:  # B열 (전화번호)
                    cell.number_format = '@'  # 문자열 형식 지정
        
        # 원본 워크북의 다른 시트 복사 (안전하게 처리)
        try:
            # 데이터만 읽기 옵션 (스타일 제외)
            original_workbook = openpyxl.load_workbook(input_file, read_only=True, data_only=True)
            original_sheets = original_workbook.sheetnames
            
            # 첫 번째 시트 이외의 다른 시트가 있는 경우에만 복사 시도
            if len(original_sheets) > 1:
                print("추가 시트 복사 중...")
                for sheet_name in original_sheets[1:]:
                    source_sheet = original_workbook[sheet_name]
                    target_sheet = new_workbook.create_sheet(title=sheet_name)
                    
                    # 데이터만 복사 (스타일 제외)
                    for row in source_sheet.iter_rows(values_only=True):
                        target_sheet.append(row)
        except Exception as e:
            print(f"추가 시트 복사 중 오류 발생: {str(e)}")
            print("주요 데이터는 처리되었으며, 추가 시트 복사는 건너뜁니다.")
        
        # 결과 파일 저장
        try:
            new_workbook.save(output_file)
            print(f"파일 변환 완료: {output_file}")
            return True
        except Exception as e:
            print(f"파일 저장 중 오류 발생: {str(e)}")
            return False
            
    except Exception as e:
        print(f"파일 변환 중 오류 발생: {str(e)}")
        traceback.print_exc()
        return False


# 예외 목록을 생성하는 함수
def create_exception_list(exception_list, output_file="exceptions.json"):
    """
    예외 상품명 목록을 JSON 파일로 저장하는 함수
    
    Args:
        exception_list (list): 예외 처리할 상품명 목록
        output_file (str): 출력할 JSON 파일 경로 (기본값: exceptions.json)
    """
    try:
        data = {"exception_products": exception_list}
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        print(f"예외 목록 파일 생성 완료: {output_file}")
        return True
    except Exception as e:
        print(f"예외 목록 파일 생성 중 오류 발생: {str(e)}")
        return False


if __name__ == "__main__":
    # 입력 파일 존재 확인
    if not os.path.exists("input.xlsx"):
        print("오류: 'input.xlsx' 파일이 현재 폴더에 존재하지 않습니다.")
    else:
        # 예외 처리 예시 (필요시 사용)
        # exception_products = ["특별 상품명1", "합치지 않을 상품명2"]
        # create_exception_list(exception_products)
        
        # 파일 변환 실행
        transform_excel_file()